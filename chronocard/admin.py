from datetime import timedelta, datetime

from django.contrib import admin
from django.utils.translation import gettext_lazy as _
from django.contrib.auth.forms import UserChangeForm, UserCreationForm
from django.contrib.auth.admin import UserAdmin
from django.http import HttpResponse

from .models import User, Event, EventShift, Location, EventUser, Checkin

# Register your models here.


def export_xlsx(modeladmin, request, queryset):
    def daterange(start_date, end_date):
        for n in range(int((end_date - start_date).days + 1)):
            yield start_date + timedelta(n)

    def _convert_seconds_to_str(seconds):
        in_seconds = seconds
        if not isinstance(seconds, int):
            try:
                in_seconds = int(seconds)
            except ValueError:
                in_seconds = int(float(seconds))

        h = in_seconds // 3600
        m = (in_seconds % 3600) // 60
        sec = (in_seconds % 3600) % 60
        result = '{}:{:0>2d}:{:0>2d}'.format(h, m, sec)
        return result

    import openpyxl
    from openpyxl.utils import get_column_letter
    ts = datetime.now().strftime("%Y-%m-%dT%H:%M:%S")
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=event_export_' + ts + '.xlsx'
    wb = openpyxl.Workbook()
    std = wb['Sheet']
    wb.remove(std)
    for event in queryset:
        ws = wb.create_sheet(event.name)

        work_dates = []
        for single_date in daterange(event.work_start_date, event.work_end_date):
            work_dates.append(single_date)

        row_num = 0

        columns = [
            ('Handle', 20),
        ]
        for work_date in work_dates:
            columns.append((str(work_date), 20))
        columns.append(('Total', 20))

        for col_num in range(len(columns)):
            c = ws.cell(row=row_num + 1, column=col_num + 1)
            c.value = columns[col_num][0]
            # c.style.font = Font(bold=True)
            # set column width
            ws.column_dimensions[get_column_letter(col_num+1)].width = columns[col_num][1]

        event_users = event.eventuser_set.all().order_by('user__handle')
        for event_user in event_users:
            row_num += 1

            data = event_user.time_report
            data['Handle'] = event_user.user.handle
            data['Total'] = event_user.total_time

            row_data = [str(data[x[0]]) if x[0] == 'Handle' else
                        _convert_seconds_to_str(data[x[0]].total_seconds()) for x in columns]

            for col_num in range(len(row_data)):
                c = ws.cell(row=row_num + 1, column=col_num + 1)
                c.value = row_data[col_num]
                # c.style.alignment.wrap_text = True

    wb.save(response)
    return response


export_xlsx.short_description = 'Export to XLSX'


class EnforcerUserChangeForm(UserChangeForm):
    class Meta(UserChangeForm.Meta):
        model = User
        fields = ('username', 'first_name', 'last_name', 'email', 'handle')


class EnforcerUserCreationForm(UserCreationForm):
    class Meta(UserCreationForm.Meta):
        model = User
        # fields = ('username', 'first_name', 'last_name', 'email', 'handle')


class EnforcerUserAdmin(UserAdmin):
    add_form = EnforcerUserCreationForm
    form = EnforcerUserChangeForm
    list_display = ('id', 'handle', 'first_name', 'last_name', 'email', 'username')
    # fields = ('id', 'handle', 'first_name', 'last_name', 'email')
    fieldsets = (
        (_('Personal info'), {'fields': ('first_name', 'last_name', 'email', 'handle')}),
        (None, {'fields': ('username', 'password')}),
        (_('Permissions'), {'fields': ('is_active', 'is_staff', 'is_superuser', 'groups',
                                       # 'user_permissions'
                                       )}),
        # (_('Important dates'), {'fields': ('last_login', 'date_joined')}),
    )
    add_fieldsets = (
        (None, {
            'classes': ('wide',),
            'fields': ('username', 'handle', 'first_name', 'last_name', 'email', 'password1', 'password2'),
        }),
    )


class CheckinInline(admin.TabularInline):
    model = Checkin
    # list_display = ('start_date', 'end_date', 'total_time')
    fields = ('start_date', 'end_date', 'total_time')
    ordering = ('start_date',)
    readonly_fields = ['total_time', ]
    extra = 1


class EnforcerEventUserAdmin(admin.ModelAdmin):
    list_display = ('id', 'get_user_handle', 'get_event_name', 'badge_id', 'total_time',)
    list_filter = ('event__name', )
    list_per_page = 100
    ordering = ('id', 'user__handle')
    inlines = [CheckinInline, ]
    # fields = ['event', 'user', 'badge_id', 'total_time']
    readonly_fields = ['total_time', ]

    def get_user_handle(self, obj):
        return obj.user.handle

    def get_event_name(self, obj):
        return obj.event.name

    get_user_handle.admin_order_field = 'user'
    get_user_handle.short_description = 'User Handle'
    get_event_name.admin_order_field = 'event'
    get_event_name.short_description = 'Event'


class EnforcerEventAdmin(admin.ModelAdmin):
    list_display = ('id', 'name', 'start_date', 'end_date')
    ordering = ('id',)
    actions = [export_xlsx]


admin.site.register(User, EnforcerUserAdmin)
admin.site.register(EventUser, EnforcerEventUserAdmin)
admin.site.register(Location)
admin.site.register(Event, EnforcerEventAdmin)
admin.site.register(EventShift)
